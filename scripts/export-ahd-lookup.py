"""Export AHD_lookup.xlsx to public/data/ahd-lookup.json"""
import json
import re
import shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "AHD_lookup.xlsx"
OUT_JSON = ROOT / "public" / "data" / "ahd-lookup.json"
LOGOS_SRC = Path(r"C:\Users\JohnSt\Desktop\RNZ\School Logos")
LOGOS_DST = ROOT / "public" / "assets" / "school-logos"
KRI_SRC = ROOT
KRI_DST = ROOT / "public" / "assets" / "vmix" / "kri"

KRI_FILES = {
    "title": "KRI_ Title 02.png",
    "lower": "KRI_ Lower Third.png",
    "draw": "KRI_ Draw.png",
    "results": "KRI_ Results.png",
    "leader": "KRI_ Leader.png",
}


def col_map(ws, code_col=0, name_col=1):
    m = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = row[code_col] if code_col < len(row) else None
        name = row[name_col] if name_col < len(row) else None
        if code and name:
            m[str(code).strip()] = str(name).strip()
    return m


def export_lookup():
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    clubs_ws = wb["Clubs"]
    events_ws = wb["Events"]

    clubs = {}
    for row in clubs_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cid = str(row[0]).strip().lower()
        clubs[cid] = {
            "id": cid,
            "name": (str(row[1]).strip() if row[1] else cid.upper()),
            "type": (str(row[2]).strip() if row[2] else ""),
            "logo": (str(row[3]).strip() if row[3] else ""),
        }

    gender = col_map(events_ws, 0, 1)
    boat = col_map(events_ws, 6, 7)
    klass = col_map(events_ws, 3, 4)
    wb.close()

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps({"clubs": clubs, "gender": gender, "class": klass, "boat": boat}, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {OUT_JSON} ({len(clubs)} clubs)")
    return clubs


def normalize_logo_filename(name):
    return re.sub(r'[<>:"/\\|?*]', "", name.strip())


def copy_logos(clubs):
    LOGOS_DST.mkdir(parents=True, exist_ok=True)
    copied = 0
    for c in clubs.values():
        logo = c.get("logo") or ""
        if not logo:
            continue
        src_name = normalize_logo_filename(logo)
        src = LOGOS_SRC / src_name
        if not src.is_file():
            alt = LOGOS_SRC / src_name.replace(" - ", " - ").upper()
            if alt.is_file():
                src = alt
            else:
                for f in LOGOS_SRC.glob("*"):
                    if f.name.lower() == src_name.lower():
                        src = f
                        break
        if not src.is_file():
            continue
        dst = LOGOS_DST / src.name
        if not dst.is_file():
            shutil.copy2(src, dst)
            copied += 1
    print(f"Copied {copied} school logos to {LOGOS_DST}")


def copy_kri_backgrounds():
    KRI_DST.mkdir(parents=True, exist_ok=True)
    for key, fname in KRI_FILES.items():
        src = KRI_SRC / fname
        if src.is_file():
            shutil.copy2(src, KRI_DST / f"{key}.png")
            print(f"  {key} <- {fname}")


if __name__ == "__main__":
    clubs = export_lookup()
    copy_logos(clubs)
    print("KRI backgrounds:")
    copy_kri_backgrounds()
